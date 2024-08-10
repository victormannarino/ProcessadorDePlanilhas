import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import subprocess
from datetime import datetime

novo_diretorio = 'Diretório'#Inserir diretório dos arquivos
os.chdir(novo_diretorio)
print("Diretório atual:", os.getcwd())



# Funções dos botões
def criar_planilha():
    
        
    novo_diretorio = 'Diretório' #Inserir diretório dos arquivos
    os.chdir(novo_diretorio)
    print("Diretório atual:", os.getcwd())

    # Carregar as planilhas
    planilha_macro = pd.ExcelFile('NMacro.xlsx')
    planilha_resumo = pd.ExcelFile('Resumo.xlsx')
    planilha_dados = pd.ExcelFile('dados.xlsx')

    # Função para remover colunas de um DataFrame
    def remover_colunas(df, colunas_a_remover):
        df = df.drop(columns=colunas_a_remover, errors='ignore')
        return df

    # Remover colunas da planilha chamada "Resumo"
    resumo_df = planilha_resumo.parse(sheet_name='Sheet1')  # Ajuste o nome da aba se necessário
    colunas_resumo = [18, 17, 14, 13, 12, 10, 5, 3, 2, 1]
    colunas_resumo = [col - 1 for col in colunas_resumo]  # Ajustar índice de 1-base para 0-base
    resumo_df = remover_colunas(resumo_df, resumo_df.columns[colunas_resumo].tolist())

    # Salvar a planilha "Resumo" atualizada
    with pd.ExcelWriter('Resumo.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        resumo_df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Remover colunas da planilha "Dados"
    dados_df = planilha_dados.parse(sheet_name='Sheet1')  # Ajuste o nome da aba se necessário
    colunas_dados = list(range(34, 16, -1)) + [15, 14, 13, 12, 11, 9, 8, 7, 6, 5, 3, 2]
    colunas_dados = [col - 1 for col in colunas_dados]  # Ajustar índice de 1-base para 0-base
    dados_df = remover_colunas(dados_df, dados_df.columns[colunas_dados].tolist())

    # Salvar a planilha "Dados" atualizada
    with pd.ExcelWriter('dados.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        dados_df.to_excel(writer, sheet_name='Sheet1', index=False)

    print("Colunas removidas e planilhas salvas com sucesso.")


    # 1. Inserir coluna 9 na planilha Resumo chamada "Quant. Materiais"
    def calcular_quant_materiais(resumo_df, dados_df):
        resumo_df.insert(8, 'Quant. Materiais', 0)
        for index, row in resumo_df.iterrows():
            valor_col6 = row[5]  # Coluna 6
            count = dados_df[dados_df.iloc[:, 0] == valor_col6].shape[0]
            resumo_df.at[index, 'Quant. Materiais'] = count
        return resumo_df

    # 2. Inserir coluna 10 na planilha Resumo chamada "Quant. Peças"
    def calcular_quant_pecas(resumo_df, dados_df):
        resumo_df.insert(9, 'Quant. Peças', 0)
        for index, row in resumo_df.iterrows():
            valor_col6 = row[5]  # Coluna 6
            soma_pecas = dados_df[dados_df.iloc[:, 0] == valor_col6].iloc[:, 2].sum()  # Coluna 3 é o índice 2
            resumo_df.at[index, 'Quant. Peças'] = soma_pecas
        return resumo_df

    # 3. Formatando a coluna 4 da planilha Dados como número e adicionar coluna "Tipo"
    def formatar_coluna_e_adicionar_tipo(dados_df, macro_df):
        dados_df.iloc[:, 3] = pd.to_numeric(dados_df.iloc[:, 3], errors='coerce')  # Formatando coluna 4 como número
        dados_df.insert(4, 'Tipo', '')  # Adicionando a coluna "Tipo"
        for index, row in dados_df.iterrows():
            valor_col4 = row[3]  # Coluna 4
            tipo = macro_df[macro_df.iloc[:, 0] == valor_col4].iloc[0, 2] if not macro_df[macro_df.iloc[:, 0] == valor_col4].empty else ''
            dados_df.at[index, 'Tipo'] = tipo
        return dados_df

    # 4. Inserir coluna 11 na planilha Resumo chamada "Categoria"
    def adicionar_categoria(resumo_df, dados_df):
        resumo_df.insert(10, 'Categoria', '')  # Adicionando a coluna "Categoria"
        for index, row in resumo_df.iterrows(): 
            valor_col6 = row[5]  # Coluna 6
            categoria = dados_df[dados_df.iloc[:, 0] == valor_col6].iloc[0, 4] if not dados_df[dados_df.iloc[:, 0] == valor_col6].empty else ''
            resumo_df.at[index, 'Categoria'] = categoria
        return resumo_df

    # 5. Atualizar coluna 8 da planilha Resumo com substituição
    def atualizar_coluna8(resumo_df):
        resumo_df.iloc[:, 7] = resumo_df.iloc[:, 7].apply(lambda x: 'BONIFICAÇÃO' if x == 0 else 'NÃO')
        return resumo_df

    # Aplicar as funções
    resumo_df = calcular_quant_materiais(planilha_resumo, planilha_dados)
    resumo_df = calcular_quant_pecas(resumo_df, planilha_dados)
    dados_df = formatar_coluna_e_adicionar_tipo(planilha_dados, planilha_macro)
    resumo_df = adicionar_categoria(resumo_df, planilha_dados)
    resumo_df = atualizar_coluna8(resumo_df)

    # Salvar as planilhas atualizadas
    with pd.ExcelWriter('Resumo.xlsx', engine='openpyxl', mode='w') as writer:
        resumo_df.to_excel(writer, sheet_name='Sheet1', index=False)

    with pd.ExcelWriter('dados.xlsx', engine='openpyxl', mode='w') as writer:
        dados_df.to_excel(writer, sheet_name='Sheet1', index=False)

    print("Colunas adicionadas e atualizações realizadas com sucesso.")

    # Função para criar a planilha Envio
    def criar_planilha_envio(resumo_df):
        # Definir a ordem das colunas para a nova planilha Envio
        colunas_envio = {
            'Código': resumo_df.columns[0],  # Coluna 1 da resumo
            'Fornecedor': resumo_df.columns[1],  # Coluna 2 da resumo
            'Data de Faturamento': resumo_df.columns[3],  # Coluna 4 da resumo
            'Data de Entrada': resumo_df.columns[4],  # Coluna 5 da resumo
            'NFs': resumo_df.columns[2],  # Coluna 3 da resumo
            'Quant. de SKUs': resumo_df.columns[8],  # Coluna 9 da resumo
            'Quant. Peças': resumo_df.columns[9],  # Coluna 10 da resumo
            'Categoria': resumo_df.columns[10], # Coluna 11 da resumo
            'BONIFICAÇÃO?': resumo_df.columns[7]   # Coluna 8 da resumo
        }

        # Reorganizar as colunas conforme a ordem especificada
        envio_df = resumo_df[list(colunas_envio.values())]
        envio_df.columns = list(colunas_envio.keys())

        # Salvar a nova planilha Envio
        with pd.ExcelWriter('Envio.xlsx', engine='openpyxl', mode='w') as writer:
            envio_df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Criar a planilha Envio
    criar_planilha_envio(planilha_resumo)
    print("Planilha Envio criada e salva com sucesso.")


    pass        

def abrir_planilha():
        
    # Caminho  do arquivo
    file_path = r"C:\Users\19751\Documents\GitHub\Conveniencia\Envio.xlsx"
    os.startfile(file_path)
         
    pass

def formatar_planilha():
    os.chdir('C:/Users/19751/Documents/GitHub/Conveniencia')   
    # Carregar o arquivo Excel
    file_path = 'envio.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet1']

    # Carregar o DataFrame
    df = pd.read_excel(file_path, sheet_name='Sheet1')

    ###################################################################
    ###################################################################

    # Substituir valores na coluna 5 pelo número 1
    df.iloc[:, 4] = 1

    # Atualizar a planilha com o DataFrame modificado
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    print("Valores na coluna 5 substituídos por 1")

    ########################################################################

    # Verificar se está escrito "BONIFICAÇÃO" na coluna 9 e agrupar com a coluna 8
    df.loc[df.iloc[:, 8] == 'BONIFICAÇÃO', df.columns[7]] += ' BONIFICAÇÃO'


    # Apagar a coluna 9
    df.drop(df.columns[8], axis=1, inplace=True)

    # Atualizar a planilha com o DF modificado
    ws.delete_cols(9)  # Apagar a coluna 9 diretamente na planilha
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    print("Texto agrupado e coluna 9 apagada com sucesso.")

    ########################################################################


    # Definir as colunas a serem somadas e as colunas a serem agrupadas
    sum_columns = [df.columns[4], df.columns[5], df.columns[6]]
    group_columns = [df.columns[0], df.columns[2], df.columns[3], df.columns[7]]

    # Agrupar por colunas 1, 3, 4 e 8, somando as colunas 5, 6 e 7 e mantendo outras colunas
    agg_dict = {col: 'first' for col in df.columns}
    for col in sum_columns:
        agg_dict[col] = 'sum'

    grouped_df = df.groupby(group_columns, as_index=False).agg(agg_dict)

    # Limpar a planilha antes de atualizar
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Atualizar a planilha com o DataFrame modificado
    for r_idx, row in enumerate(dataframe_to_rows(grouped_df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Salvar o arquivo atualizado
    print("Linhas agrupadas e somadas com sucesso, mantendo a posição das colunas.")

    # Apagar a coluna 1 (coluna A)
    ws.delete_cols(1)
    print("coluna 1 apagada")


    ##################################################################


    # Encontrar a primeira célula vazia na coluna 1
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        if ws[f'A{row}'].value is None:
            first_empty_row = row
            break
    else:
        first_empty_row = max_row + 1

    # Adicionar linha TOTAL com a soma das colunas
    ws[f'A{first_empty_row}'] = 'TOTAL'
    ws.merge_cells(start_row=first_empty_row, start_column=1, end_row=first_empty_row + 1, end_column=3)
    ws[f'D{first_empty_row}'] = f'=SUM(D1:D{first_empty_row - 1})'
    ws[f'E{first_empty_row}'] = f'=SUM(E1:E{first_empty_row - 1})'
    ws[f'F{first_empty_row}'] = f'=SUM(F1:F{first_empty_row - 1})'
    ws.merge_cells(start_row=first_empty_row, start_column=4, end_row=first_empty_row + 1, end_column=4)
    ws.merge_cells(start_row=first_empty_row, start_column=5, end_row=first_empty_row + 1, end_column=5)
    ws.merge_cells(start_row=first_empty_row, start_column=6, end_row=first_empty_row + 1, end_column=6)
    ws.merge_cells(start_row=first_empty_row, start_column=7, end_row=first_empty_row + 1, end_column=7)

    # Salvar o arquivo atualizado
    print ("TOTAL concluido")
    ###################################################################


    # Cor e fonte
    dark_blue_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    white_font = Font(color="FFFFFF", name="Aptos Display")  # Define a fonte como Arial Narrow
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Formatar cabeçalho e linha TOTAL
    for col in range(1, 8):  # Colunas A a G
        header_cell = ws.cell(row=1, column=col)
        total_cell = ws.cell(row=first_empty_row, column=col)
        header_cell.fill = dark_blue_fill
        total_cell.fill = dark_blue_fill
        header_cell.font = white_font
        total_cell.font = white_font
        header_cell.alignment = center_alignment
        total_cell.alignment = center_alignment

    # Alinhar toda a tabela
    for row in ws.iter_rows(min_row=1, max_row=first_empty_row + 1, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = center_alignment

    print ("pintou")

    # Ajustar o tamanho das colunas
    column_widths = {
        'A': 45,  # Largura da coluna 1
        'B': 18,  # Largura da coluna 2
        'C': 15, # Largura da coluna 3
        'D': 10,  # Largura da coluna 4
        'E': 15,  # Largura da coluna 5
        'F': 15,  # Largura da coluna 6
        'G': 20,   # Largura da coluna 7
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    print ("colunas ajustadas")

    for cell in ws['A']:
        cell.alignment = Alignment(horizontal='left')


    # Definir as cores
    colors = {
        'HPC - CONV': '9A57CD',
        'HPC - CONV BONIFICAÇÃO': '9A57CD',
        'HPC': 'FFC000',
        'HPC BONIFICAÇÃO': 'FFC000',
        'MED - GEN/SIM': '8DB4E2',
        'MED - GEN/SIM BONIFICAÇÃO': '8DB4E2',
        'MED - REF': '538DD5',
        'MED - REF BONIFICAÇÃO': '538DD5'
    }

    # Definir a fonte
    font_aptos_display = Font(name='Aptos Display')
    font_aptos_display_bold = Font(name='Aptos Display', bold=True)
    font_white_bold = Font(name='Aptos Display', bold=True, color='FFFFFF')

    # Obter o número máximo de linhas na planilha
    max_row = ws.max_row
    max_column = ws.max_column

    # Aplicar a fonte Aptos Display a todas as colunas de A a G
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=7):
        for cell in row:
            cell.font = font_aptos_display

    # Aplicar as cores às células na coluna G
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=7)  # Coluna G é a coluna 7
        if cell.value in colors:
            fill = PatternFill(start_color=colors[cell.value], end_color=colors[cell.value], fill_type='solid')
            cell.fill = fill
        cell.font = font_aptos_display

    # Aplicar negrito e cor branca a todas as células da linha 1
    for cell in ws[1]:
        cell.font = Font(name='Aptos Display', bold=True, color='FFFFFF')

    # Aplicar negrito às células na coluna G mantendo a cor original
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=7)  
        if cell.value in colors:
            fill = PatternFill(start_color=colors[cell.value], end_color=colors[cell.value], fill_type='solid')
            cell.fill = fill
    
        # Aplicar negrito mantendo a cor original da fonte
        cell.font = Font(name='Aptos Display', bold=True, color=cell.font.color or '000000')


    # Aplicar negrito e cor branca a todas as células da linha 1
    for cell in ws[1]:
        cell.font = Font(name='Aptos Display', bold=True, color='FFFFFF')

    # Atualizar a fonte da célula da linha TOTAL para negrito, cor branca e tamanho 20
    total_font = Font(name='Aptos Display', bold=True, color='FFFFFF', size=14)

    # Atualizar a célula da linha TOTAL
    for col in range(1, 8):  # Colunas A a G
        total_cell = ws.cell(row=first_empty_row, column=col)
        total_cell.font = total_font
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar o alinhamento das células da linha TOTAL
    total_row = ws[first_empty_row]
    for cell in total_row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Certifique-se de que a célula 'TOTAL' está centralizada na linha
    ws[f'A{first_empty_row}'].alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file_path)
    print("Planilha formatada e salva com sucesso.")

    pass

#########################################
####### Criar a interface gráfica #######
#########################################

root = tk.Tk()
root.title("Bem-Vindo")

# Cores da interface
dark_purple = "#4B0082"
light_purple = "#D8BFD8"
root.configure(bg=dark_purple)

# Adicionar a mensagem de boas-vindas
welcome_label = tk.Label(root, text="Bem-Vindo", bg=dark_purple, fg="white", font=("Helvetica", 16))
welcome_label.pack(pady=10)

# Configurar o frame com cor de fundo clara
frame = tk.Frame(root, bg=light_purple)
frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Adicionar os botões
create_button = tk.Button(frame, text="Criar a Planilha Nova", command=criar_planilha, bg="#8A2BE2", fg="white", font=("Helvetica", 12))
create_button.pack(pady=5)

open_button = tk.Button(frame, text="Abrir Planilha", command=abrir_planilha, bg="#8A2BE2", fg="white", font=("Helvetica", 12))
open_button.pack(pady=5)

format_button = tk.Button(frame, text="Formatação", command=formatar_planilha, bg="#8A2BE2", fg="white", font=("Helvetica", 12))
format_button.pack(pady=20)

root.mainloop()
